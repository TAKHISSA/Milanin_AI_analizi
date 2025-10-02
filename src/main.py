import pandas as pd
from config.settings import settings
from src.data_processor import DataProcessor
from src.chat_analyzer import ChatAnalyzer
from src.report_generator import ReportGenerator
from utils.helpers import save_to_txt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def clean_excel_data(df):
    """Excel'de geçersiz karakterleri temizler"""
    illegal_chars = ['\x00', '\x01', '\x02', '\x03', '\x04', '\x05', '\x06', '\x07', 
                    '\x08', '\x0b', '\x0c', '\x0e', '\x0f', '\x10', '\x11', '\x12', 
                    '\x13', '\x14', '\x15', '\x16', '\x17', '\x18', '\x19', '\x1a', 
                    '\x1b', '\x1c', '\x1d', '\x1e', '\x1f', '\x7f']
    
    for column in df.columns:
        if df[column].dtype == 'object':
            df[column] = df[column].astype(str).apply(
                lambda x: ''.join(c for c in x if c not in illegal_chars) if isinstance(x, str) else x
            )
    
    return df

def add_excel_formatting(df, output_path):
    """Excel'e akıllı renklendirme ve özet ekler"""
    try:
        # Workbook'u aç
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Renk tanımlamaları
        correct_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')  # Açık yeşil - DOĞRU
        incorrect_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')  # Açık kırmızı - YANLIŞ
        header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')  # Mavi başlık
        header_font = Font(color='FFFFFF', bold=True, size=12)
        summary_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Altın sarı - özet
        summary_font = Font(bold=True, size=11)
        
        # Başlık satırını formatla
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sütun indekslerini bul
        column_mapping = {}
        for idx, cell in enumerate(ws[1], 1):
            column_mapping[cell.value] = idx
        
        total_rows = 0
        correct_rows = 0
        
        # Her satır için renklendirme
        for row in range(2, ws.max_row + 1):
            total_rows += 1
            row_correct = True
            
            # Karşılaştırma yapılacak alanlar
            comparison_fields = [
                ('yanıt_durumu', 'gercek_yanit_durumu'),
                ('sentiment', 'gercek_sentiment'),
                ('tür', 'gercek_tur'),
                ('intent', 'gercek_intent')
            ]
            
            for ai_field, truth_field in comparison_fields:
                if ai_field in column_mapping and truth_field in column_mapping:
                    ai_col = column_mapping[ai_field]
                    truth_col = column_mapping[truth_field]
                    
                    ai_value = ws.cell(row=row, column=ai_col).value
                    truth_value = ws.cell(row=row, column=truth_col).value
                    
                    # Değerleri karşılaştır (case-insensitive ve boşlukları yok say)
                    if (ai_value and truth_value and 
                        str(ai_value).lower().strip() == str(truth_value).lower().strip()):
                        ws.cell(row=row, column=ai_col).fill = correct_fill
                    else:
                        ws.cell(row=row, column=ai_col).fill = incorrect_fill
                        row_correct = False
            
            if row_correct:
                correct_rows += 1
        
        # Özet satırları ekle
        summary_row = ws.max_row + 3
        
        # Özet başlığı
        ws.merge_cells(f'A{summary_row}:B{summary_row}')
        ws.cell(row=summary_row, column=1).value = "📊 PERFORMANS ÖZETİ"
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=14, color='2E86AB')
        ws.cell(row=summary_row, column=1).fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
        ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal='center')
        
        # Özet verileri
        summary_data = [
            ("🔢 Toplam Sohbet", total_rows),
            ("✅ Tam Doğru Tahmin", correct_rows),
            ("⚠️ Kısmi Doğru Tahmin", total_rows - correct_rows),
            ("🎯 Tam Doğruluk Oranı", f"%{(correct_rows/total_rows)*100:.1f}" if total_rows > 0 else "%0.0")
        ]
        
        for i, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=summary_row + i, column=1).value = label
            ws.cell(row=summary_row + i, column=1).font = summary_font
            ws.cell(row=summary_row + i, column=2).value = value
            ws.cell(row=summary_row + i, column=2).font = summary_font
        
        # Renk Açıklaması
        legend_row = summary_row + len(summary_data) + 2
        ws.cell(row=legend_row, column=1).value = "🎨 RENK AÇIKLAMASI"
        ws.cell(row=legend_row, column=1).font = Font(bold=True, size=12, color='2E86AB')
        
        legend_data = [
            ("🟢 Yeşil", "Doğru Tahmin"),
            ("🔴 Kırmızı", "Yanlış Tahmin")
        ]
        
        for i, (color, meaning) in enumerate(legend_data, 1):
            ws.cell(row=legend_row + i, column=1).value = color
            ws.cell(row=legend_row + i, column=2).value = meaning
        
        # Sütun genişliklerini ayarla
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        print("🎨 Excel formatlama ve renklendirme uygulandı")
        
    except Exception as e:
        print(f"❌ Excel formatlama hatası: {e}")

def main():
    print("🤖 Sohbet Analiz Sistemi Başlatılıyor...")
    print("=" * 60)
    
    # Veri işlemciyi başlat
    print("📂 Veri yükleniyor...")
    processor = DataProcessor(settings.DATA_FILE)
    chats = processor.get_all_chats()
    ground_truth = processor.get_ground_truth_labels()
    
    if not chats:
        print("❌ Sohbet verisi bulunamadı!")
        return
    
    print(f"✅ {len(chats)} sohbet bulundu.")
    
    # Analizörü başlat
    analyzer = ChatAnalyzer()
    
    # Tüm sohbetleri analiz et
    print("\n🔍 Sohbetler analiz ediliyor...")
    results = analyzer.analyze_all_chats(chats)
    
    # Excel'e kaydet
    print("\n💾 Excel raporu oluşturuluyor...")
    df = pd.DataFrame(results)
    
    # Gereksiz sütunları kaldır
    if 'raw_chat' in df.columns:
        df = df.drop('raw_chat', axis=1)
    
    # Excel için veriyi temizle
    df = clean_excel_data(df)
    
    # Sütun sırasını düzenle
    column_order = [
        'sohbet_id', 'sohbet_baslangic', 'sohbet_bitis', 'toplam_sure',
        'yanıt_durumu', 'sentiment', 'tür', 'intent', 'intent_detay',
        'gercek_yanit_durumu', 'gercek_sentiment', 'gercek_tur', 
        'gercek_intent', 'gercek_intent_detay'
    ]
    
    # Sadece mevcut sütunları al
    existing_columns = [col for col in column_order if col in df.columns]
    df = df[existing_columns]
    
    # Excel dosyasını kaydet
    try:
        df.to_excel(settings.OUTPUT_EXCEL, index=False, engine='openpyxl')
        print(f"✅ Excel raporu kaydedildi: {settings.OUTPUT_EXCEL}")
        
        # Formatlama ekle
        add_excel_formatting(df, settings.OUTPUT_EXCEL)
        
    except Exception as e:
        print(f"❌ Excel kaydetme hatası: {e}")
        # Hata durumunda CSV olarak kaydet
        csv_path = settings.OUTPUT_EXCEL.replace('.xlsx', '.csv')
        df.to_csv(csv_path, index=False, encoding='utf-8')
        print(f"✅ CSV olarak kaydedildi: {csv_path}")
    
    # Raporları oluştur ve kaydet
    print("\n📊 Raporlar oluşturuluyor...")
    report_generator = ReportGenerator()
    
    # Doğruluk raporu
    accuracy_results = report_generator.calculate_accuracy(results, ground_truth)
    accuracy_report = report_generator.generate_accuracy_report(accuracy_results)
    save_to_txt("outputs/doğruluk_raporu.txt", accuracy_report)
    
    # Diğer raporlar
    output_paths = {
        'swot': settings.OUTPUT_SWOT,
        'recommendations': settings.OUTPUT_RECOMMENDATIONS,
        'demand_summary': settings.OUTPUT_DEMAND_SUMMARY
    }
    
    reports = report_generator.save_all_reports(results, ground_truth, output_paths)
    
    print("=" * 60)
    print("🎉 ANALİZ TAMAMLANDI")
    print("=" * 60)
    print(f"📊 Excel raporu: {settings.OUTPUT_EXCEL}")
    print(f"📈 Doğruluk raporu: outputs/doğruluk_raporu.txt")
    print(f"🔍 SWOT analizi: {settings.OUTPUT_SWOT}")
    print(f"💡 Öneriler: {settings.OUTPUT_RECOMMENDATIONS}")
    print(f"📋 Talep özeti: {settings.OUTPUT_DEMAND_SUMMARY}")
    
    # Doğruluk sonuçlarını ekrana yazdır
    print("\n📊 DOĞRULUK SONUÇLARI:")
    for field, accuracy in accuracy_results.items():
        print(f"   {field}: %{accuracy}")

if __name__ == "__main__":
    main()
